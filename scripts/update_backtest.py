#!/usr/bin/env python3
"""
Fully automated backtest update. Checks SEC EDGAR for a new N-PORT
filing; if found for a review that has pending predictions in the
workbook's Predictions sheet, scores them and rolls the constituent
baseline forward. If nothing new is found, exits quietly - this is
expected most months.

Reads pending predictions directly from the workbook's Predictions
sheet (the single source of truth) rather than a separate tracking
file - avoids two copies of the same state drifting out of sync.

USAGE (no arguments needed - infers the target quarter automatically)
    python scripts/update_backtest.py
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from ftse_predictor.fetchers import find_nport_filing
from ftse_predictor.report import update_outcomes, apply_formatting
from ftse_predictor.baseline import _load_constituent_baseline

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
OUT_DIR = Path(__file__).resolve().parent.parent / "outputs"
BASELINE_PATH = DATA_DIR / "constituents_baseline.csv"
WORKBOOK_PATH = OUT_DIR / "rebalancing_predictions.xlsx"


def _candidate_quarter_ends():
    """The last 3 quarter-end dates - covers the realistic window a
    filing could have just landed for."""
    today = pd.Timestamp.now()
    ends = []
    for months_back in (0, 3, 6):
        d = today - pd.DateOffset(months=months_back)
        q_end_month = ((d.month - 1) // 3) * 3 + 3
        year = d.year if q_end_month <= 12 else d.year + 1
        q_end_month = q_end_month if q_end_month <= 12 else q_end_month - 12
        last_day = pd.Timestamp(year=year, month=q_end_month, day=1) + pd.offsets.MonthEnd(0)
        ends.append(last_day)
    return sorted(set(ends))


def _has_pending_prediction(target_date):
    """Checks the workbook's Predictions sheet directly - the single
    source of truth for what's pending, rather than a separate file
    that could fall out of sync with it."""
    if not WORKBOOK_PATH.exists() or "Predictions" not in load_workbook(
            WORKBOOK_PATH, read_only=True).sheetnames:
        return False
    wb = load_workbook(WORKBOOK_PATH, read_only=True)
    ws = wb["Predictions"]
    target_str = str(pd.Timestamp(target_date).date())
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(pd.Timestamp(row[0]).date()) == target_str:
            return True
    return False


def main():
    universe_path = DATA_DIR / "eligible_universe.csv"
    if not universe_path.exists():
        print(f"ERROR: {universe_path} not found - cannot proceed.")
        sys.exit(1)
    uni = pd.read_csv(universe_path)
    uni["isin_norm"] = uni["ISIN Code"].astype(str).str.strip().str.upper()
    isin_lookup = dict(zip(uni["Symbol"], uni["isin_norm"]))

    for target in _candidate_quarter_ends():
        target_str = str(target.date())
        print(f"Checking for N-PORT filing covering {target_str}...")
        holdings = find_nport_filing(target_str)
        if holdings is None:
            print(f"  not found yet")
            continue

        print(f"  FOUND - {len(holdings)} holdings")
        holdings["isin_norm"] = holdings["isin"].astype(str).str.strip().str.upper()
        new_isins = set(holdings["isin_norm"])

        if BASELINE_PATH.exists():
            # THE FIX: the old baseline may be in EITHER format (simple
            # isin/name, or a raw FTSE export needing fuzzy name
            # matching) - use the same shared loader run_prediction.py
            # uses, rather than assuming a plain 'isin' column always
            # exists. That assumption crashed the first live run
            # ("KeyError: 'isin'") on an otherwise-successful SEC EDGAR
            # pull, against a baseline that was still in raw FTSE
            # export format.
            old_isins, _, _ = _load_constituent_baseline(str(BASELINE_PATH), uni)
        else:
            old_isins = set()

        raw_adds = new_isins - old_isins
        raw_removes = old_isins - new_isins
        print(f"  raw adds: {len(raw_adds)}, raw removes: {len(raw_removes)}")

        if _has_pending_prediction(target):
            stats = update_outcomes(str(WORKBOOK_PATH), target, raw_adds,
                                    raw_removes, isin_lookup)
            apply_formatting(str(WORKBOOK_PATH))
            print(f"  scored and moved from Predictions to Adds/Removes "
                  f"Detail: {stats}")
        else:
            print(f"  no pending prediction found in the Predictions sheet "
                  f"for {target_str} - updating baseline only, nothing to score")

        holdings[["name", "isin"]].assign(as_of_date=target.date()).to_csv(
            BASELINE_PATH, index=False)
        print(f"  baseline rolled forward to {target_str}")


if __name__ == "__main__":
    main()
