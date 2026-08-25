"""
The permanent workbook. Two distinct kinds of rows are kept
deliberately separate:

    Predictions        PENDING predictions - not yet scored. Lives in
                       its own sheet, never mixed with real history.
    Adds/Removes Detail HISTORICAL, SCORED rows only. Every row here
                       has a definite Outcome (Caught/Missed/False
                       Positive) - if it doesn't have one yet, it does
                       not belong in this sheet.

When real N-PORT data arrives for a quarter that was predicted, the
matching rows are DELETED from Predictions and INSERTED (properly
scored) into Adds/Removes Detail - a move, not an edit-in-place.

SCHEMA (must match exactly - reading/writing an existing file with a
fixed column layout):
    Summary:         Review Date, Quarter Type, Adds Predicted,
                     Adds Actual, Adds Caught, Add Hit Rate %,
                     Removes Predicted, Removes Actual, Removes Caught
    Predictions:     Review Date, Quarter Type, Type (Add/Remove),
                     Symbol, Company, Rank, Cum. Coverage %,
                     Months Listed, Market Cap (Cr), Confidence,
                     Confidence Depth %
    Adds Detail:     Review Date, Quarter Type, Symbol, Company,
                     Predicted, Actually Added, Outcome, Rank,
                     Cum. Coverage %, Months Listed, Market Cap (Cr),
                     Confidence, Confidence Depth %
    Removes Detail:  Review Date, Quarter Type, Symbol, Company,
                     Predicted, Actually Removed, Outcome, Rank,
                     Cum. Coverage %, Market Cap (Cr), Confidence,
                     Confidence Depth %
"""

import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import CONFIG

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
WARNING_FILL = PatternFill("solid", fgColor="FF0000")
WARNING_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
CONF_FILLS = {"High": PatternFill("solid", fgColor="C6EFCE"),
              "Elevated": PatternFill("solid", fgColor="FFEB9C"),
              "Watch": PatternFill("solid", fgColor="F2F2F2")}
OUTCOME_FILLS = {"Caught": PatternFill("solid", fgColor="C6EFCE"),
                 "Missed": PatternFill("solid", fgColor="FFC7CE"),
                 "False Positive": PatternFill("solid", fgColor="FFEB9C")}

_QUARTER_TYPE_LABEL = {
    "general_review": "Mar/Sep (general review)",
    "fast_entry": "Jun/Dec (fast entry)",
}

_PREDICTIONS_HEADERS = ["Review Date", "Quarter Type", "Type", "Symbol",
                        "Company", "Rank", "Cum. Coverage %", "Months Listed",
                        "Market Cap (Cr)", "Confidence", "Confidence Depth %"]


def _add_confidence(row, quarter_type):
    if quarter_type == "general_review":
        depth = (CONFIG.entry_coverage - row["cum_coverage"]) / CONFIG.entry_coverage * 100
    else:
        mcap_cr = row["full_mcap"] / 1e7
        depth = min((mcap_cr / CONFIG.fast_entry_min_mcap_cr - 1) * 100, 100)
    tier = "High" if depth >= 60 else "Elevated" if depth >= 25 else "Watch"
    return tier, round(depth, 1)


def _remove_confidence(cum_coverage):
    depth = (cum_coverage - CONFIG.exit_coverage) / (100 - CONFIG.exit_coverage) * 100
    tier = "High" if depth >= 60 else "Elevated" if depth >= 25 else "Watch"
    return tier, round(depth, 1)


def _ensure_predictions_sheet(wb):
    if "Predictions" in wb.sheetnames:
        return wb["Predictions"]
    ws = wb.create_sheet("Predictions", index=1)   # right after Summary
    ws.append(_PREDICTIONS_HEADERS)
    return ws


def append_prediction(workbook_path, result):
    """Adds this review's predictions to the PENDING Predictions sheet
    only - never touches Adds/Removes Detail. If a pending prediction
    already exists for this review date (a re-run), it is replaced."""
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(
            f"{workbook_path} not found. This workbook should have been "
            f"seeded with real historical data before automation started - "
            f"see README. Refusing to silently create a blank one, since "
            f"that would lose the existing backtest history.")

    wb = load_workbook(workbook_path)
    pred_ws = _ensure_predictions_sheet(wb)
    _delete_rows_matching(pred_ws, col_index=1, value=result.review_date)

    label = _QUARTER_TYPE_LABEL[result.quarter_type]

    for _, r in result.adds.iterrows():
        tier, depth = _add_confidence(r, result.quarter_type)
        pred_ws.append([
            result.review_date, label, "Add", r["symbol"], r["name"],
            int(r["rank"]), round(r["cum_coverage"], 1),
            round(r["months_listed"], 1) if pd.notna(r.get("months_listed")) else None,
            round(r["full_mcap"] / 1e7, 0), tier, depth])

    for _, r in result.removes.iterrows():
        tier, depth = _remove_confidence(r["cum_coverage"])
        pred_ws.append([
            result.review_date, label, "Remove", r["symbol"], r["name"],
            int(r["rank"]), round(r["cum_coverage"], 1), None,
            round(r["full_mcap"] / 1e7, 0), tier, depth])

    summary_ws = wb["Summary"]
    _delete_rows_matching(summary_ws, col_index=1, value=result.review_date)
    summary_ws.append([result.review_date, label, len(result.adds), None, None,
                       None, len(result.removes), None, None])

    wb.save(workbook_path)


def update_outcomes(workbook_path, review_date, raw_add_isins, raw_remove_isins,
                    isin_lookup):
    """Called once real N-PORT data is available for a review that was
    predicted. MOVES the matching rows: deletes them from Predictions,
    inserts properly-scored rows into Adds/Removes Detail. Also adds
    genuinely missed real changes directly into Adds/Removes Detail as
    'Missed' rows, since those were never in Predictions to begin with."""
    wb = load_workbook(workbook_path)
    review_str = str(pd.Timestamp(review_date).date())
    pred_ws = _ensure_predictions_sheet(wb)

    stats = {"adds_caught": 0, "adds_missed": 0, "adds_false_positive": 0,
            "removes_caught": 0}

    pred_headers = [c.value for c in pred_ws[1]]
    date_i = pred_headers.index("Review Date")
    type_i = pred_headers.index("Type")
    sym_i = pred_headers.index("Symbol")
    company_i = pred_headers.index("Company")
    rank_i = pred_headers.index("Rank")
    cov_i = pred_headers.index("Cum. Coverage %")
    months_i = pred_headers.index("Months Listed")
    mcap_i = pred_headers.index("Market Cap (Cr)")
    conf_i = pred_headers.index("Confidence")
    depth_i = pred_headers.index("Confidence Depth %")

    pending_this_date, rows_to_delete = [], []
    for row in pred_ws.iter_rows(min_row=2):
        if row[date_i].value is None:
            continue
        if str(pd.Timestamp(row[date_i].value).date()) != review_str:
            continue
        pending_this_date.append(row)
        rows_to_delete.append(row[0].row)

    predicted_add_isins, predicted_remove_isins = set(), set()
    add_out_rows, remove_out_rows = [], []

    for row in pending_this_date:
        symbol = row[sym_i].value
        isin = isin_lookup.get(symbol)
        kind = row[type_i].value
        base = [row[date_i].value, row[1].value, symbol, row[company_i].value,
               True, None, None, row[rank_i].value, row[cov_i].value]

        if kind == "Add":
            predicted_add_isins.add(isin)
            was_actual = isin in raw_add_isins if isin else False
            outcome = "Caught" if was_actual else "False Positive"
            full_row = base[:5] + [was_actual, outcome, row[rank_i].value,
                                   row[cov_i].value, row[months_i].value,
                                   row[mcap_i].value, row[conf_i].value,
                                   row[depth_i].value]
            add_out_rows.append(full_row)
            stats["adds_caught" if was_actual else "adds_false_positive"] += 1
        else:
            predicted_remove_isins.add(isin)
            was_actual = isin in raw_remove_isins if isin else False
            outcome = "Caught" if was_actual else "False Positive"
            full_row = base[:5] + [was_actual, outcome, row[rank_i].value,
                                   row[cov_i].value, row[mcap_i].value,
                                   row[conf_i].value, row[depth_i].value]
            remove_out_rows.append(full_row)
            if was_actual:
                stats["removes_caught"] += 1

    for r in reversed(rows_to_delete):
        pred_ws.delete_rows(r)

    for r in add_out_rows:
        wb["Adds Detail"].append(r)
    for r in remove_out_rows:
        wb["Removes Detail"].append(r)

    missed_add_isins = raw_add_isins - predicted_add_isins
    for isin in missed_add_isins:
        wb["Adds Detail"].append([pd.Timestamp(review_date), None, None, None,
                                  False, True, "Missed", None, None, None,
                                  None, "N/A", None])
        stats["adds_missed"] += 1

    summary_ws = wb["Summary"]
    headers = [c.value for c in summary_ws[1]]
    date_i2 = headers.index("Review Date")
    for row in summary_ws.iter_rows(min_row=2):
        if row[date_i2].value and str(pd.Timestamp(row[date_i2].value).date()) == review_str:
            row[headers.index("Adds Actual")].value = \
                stats["adds_caught"] + stats["adds_missed"]
            row[headers.index("Adds Caught")].value = stats["adds_caught"]
            predicted = row[headers.index("Adds Predicted")].value or 0
            row[headers.index("Add Hit Rate %")].value = \
                round(stats["adds_caught"] / predicted * 100, 1) if predicted else None
            row[headers.index("Removes Actual")].value = len(raw_remove_isins)
            row[headers.index("Removes Caught")].value = stats["removes_caught"]
            break

    wb.save(workbook_path)
    return stats


def write_health_warning(workbook_path, health_log):
    """If any data fell back to last-known-good this run, write an
    unmissable warning at the top of the Summary sheet - big, red,
    plain language, designed to be obvious with no technical
    background needed to understand it."""
    if not health_log.has_issues:
        return
    wb = load_workbook(workbook_path)
    ws = wb["Summary"]
    ws.insert_rows(1, amount=2)
    ws["A1"] = ("WARNING: some data could not be refreshed this run and "
               "OLDER data was used instead. This prediction may be less "
               "reliable than usual. Details below.")
    ws["A1"].fill = WARNING_FILL
    ws["A1"].font = WARNING_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = " | ".join(health_log.summary_lines())
    ws["A2"].font = Font(name=FONT_NAME, italic=True, color="CC0000")
    ws.merge_cells("A2:I2")
    wb.save(workbook_path)


def _delete_rows_matching(ws, col_index, value):
    value_str = str(pd.Timestamp(value).date())
    to_delete = []
    for row in ws.iter_rows(min_row=2):
        cell = row[col_index - 1]
        if cell.value and str(pd.Timestamp(cell.value).date()) == value_str:
            to_delete.append(cell.row)
    for r in reversed(to_delete):
        ws.delete_rows(r)


def apply_formatting(workbook_path):
    """Reapplies consistent formatting after appends - openpyxl does not
    preserve conditional formatting across programmatic row inserts, so
    this is re-run at the end of every update."""
    wb = load_workbook(workbook_path)
    for name in ("Summary", "Predictions", "Adds Detail", "Removes Detail"):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        header_row = 1
        for r in range(1, 4):
            if ws.cell(row=r, column=1).value in ("Review Date",):
                header_row = r
                break
        for cell in ws[header_row]:
            if cell.value:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = f"A{header_row + 1}"

        headers = [c.value for c in ws[header_row]]
        oi = headers.index("Outcome") + 1 if "Outcome" in headers else None
        ci = headers.index("Confidence") + 1 if "Confidence" in headers else None
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.font = Font(name=FONT_NAME)
            if oi and row[oi - 1].value in OUTCOME_FILLS:
                row[oi - 1].fill = OUTCOME_FILLS[row[oi - 1].value]
            if ci and row[ci - 1].value in CONF_FILLS:
                row[ci - 1].fill = CONF_FILLS[row[ci - 1].value]

        for col in ws.columns:
            vals = [c.value for c in col if c.value is not None]
            if not vals:
                continue
            ln = max(len(str(v)) for v in vals)
            ws.column_dimensions[get_column_letter(col[0].column)].width = \
                min(max(ln + 2, 11), 50)
    wb.save(workbook_path)
